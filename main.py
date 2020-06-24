from textblob import TextBlob
import sys, tweepy
import matplotlib.pyplot as plt
from openpyxl import *

from twitterAPIKeys import CONSUMER_KEY, CONSUMER_SECRET, ACCESS_TOKEN, ACCESS_TOKEN_SECRET

filePath = "C:/Users/Waign/Desktop/Sentiment.xlsx"
wb = load_workbook(filePath)
ws = wb["Tabelle1"]

def percentage(part, whole):
    return 100 * float(part)/float(whole)



consumerKey = CONSUMER_KEY
consumerSecret = CONSUMER_SECRET
accessToken = ACCESS_TOKEN
accessTokenSecret = ACCESS_TOKEN_SECRET

#print(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_TOKEN, ACCESS_TOKEN_SECRET)

auth = tweepy.OAuthHandler(consumer_key=consumerKey, consumer_secret=consumerSecret)
auth.set_access_token(accessToken, accessTokenSecret)
api = tweepy.API(auth)

print("Connected to Twitter API")

#searchTerm = input("Enter keyword/hashtag to search about: ")
#noOfSearchTerms = int(input("Enter how many tweets to analyze: "))

searchTerm = "Warzone"
noOfSearchTerms = 500



tweets = tweepy.Cursor(api.search, q=searchTerm, lang="en", tweet_mode="extended").items(noOfSearchTerms)

positive = 0
negative = 0
neutral = 0
polarity = 0

x = 1

for tweet in tweets:
    #print(tweet.full_text)
    #print("")
    analysis = TextBlob(tweet.full_text)
    polarity += analysis.sentiment.polarity

    if(analysis.sentiment.polarity == 0):
        neutral += 1
    elif(analysis.sentiment.polarity < 0.00):
        negative += 1
    elif(analysis.sentiment.polarity > 0.00):
        positive += 1


    wcell1 = ws.cell(x, 1)
    wcell2 = ws.cell(x, 2)
    wcell1.value = tweet.full_text
    wcell2.value = analysis.sentiment.polarity
    x += 1


wb.save(filePath)

positive = percentage(positive, noOfSearchTerms)
negative = percentage(negative, noOfSearchTerms)
neutral = percentage(neutral, noOfSearchTerms)

positive = format(positive, ".2f")
neutral = format(neutral, ".2f")
negative = format(negative, ".2f")

print("How peaple are reacting on " + searchTerm + " by analyzing " + str(noOfSearchTerms) + " Tweets.")

if(polarity == 0):
    print("Neutral")
if(polarity < 0):
    print("Negative")
if(polarity > 0):
    print("Positive")

labels = ["Positive ["+str(positive)+"%]", "Neutral ["+str(neutral)+"%]", "Negative ["+str(negative)+"%]"]
sizes = [positive, neutral, negative]
colors = ["yellowgreen", "gold", "red"]
patches, texts = plt.pie(sizes, colors=colors, startangle=90)
plt.legend(patches, labels, loc="best")
plt.title(("How peaple are reacting on " + searchTerm + " by analyzing " + str(noOfSearchTerms) + " Tweets."))
plt.axis("equal")
plt.tight_layout()
plt.show()